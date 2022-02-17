import pymysql
import pandas as pd
import copy
import openpyxl
from openpyxl.styles import PatternFill


class MysqlQuest:
    def __init__(self, host='localhost', user='root', password='root', port=3306):
        self.db = pymysql.connect(host='localhost', user='root', password='root', port=3306)
        self.host = host
        self.user = user
        self.password = password
        self.port = port
        self.cursor = []
        self.database = []
        self.database_like = []
        self.table = []
        self.table_like = []
        self.colum = []
        self.colum_like = []

        # init important
        self.get_db()
        self.get_cursor()
        self.get_table()
        self.get_colum()

    def get_db(self):
        if not self.database:
            sql_setns = "show databases"
            self.cursor.append(self.db.cursor())
            self.cursor[0].execute(sql_setns)
            results_database = self.cursor[0].fetchall()
            for i in results_database:
                if i[0] != 'information_schema' and i[0] != 'performance_schema':
                    self.database.append(i[0])
            self.cursor = []

    def get_cursor(self):
        if not self.cursor:
            for i in self.database:
                j = pymysql.connect(host=self.host, user=self.user, password=self.password, port=self.port, database=i)
                j = j.cursor()
                self.cursor.append(j)

    def get_table(self):
        table_row = []
        if not self.table:
            for i, _ in enumerate(self.database):
                sql_setns = "show tables"
                self.cursor[i].execute(sql_setns)
                results_table = self.cursor[i].fetchall()
                for j in results_table:
                    table_row.append(j[0])
                self.table.append(table_row)
                table_row = []

    def get_colum(self):
        colum_row = []
        colum_columns = []
        colum_temp = []
        # for id in self.colum:
        for i, _ in enumerate(self.database):
            for j in self.table[i]:
                # sql_setns = "show columns from " + j + "like '"+str(id)+"%'"
                sql_setns = "show columns from " + j
                self.cursor[i].execute(sql_setns)
                results_colum = self.cursor[i].fetchall()
                for k in results_colum:
                    colum_temp.append(k[0])
                colum_row.append(colum_temp)
                colum_temp = []
            colum_columns.append(colum_row)
            colum_row = []
        self.colum = colum_columns

    def do_real_worlk(self, colum=None, value=None):
        # database = cursor
        if colum is None:
            colum = []
        if value is None:
            value = []
        str1 = self.creat_sql_str(self.cursor, self.table, colum, value)
        df_list = self.do_sql_from_str(str1)
        for df in df_list:
            print(df)

    @staticmethod
    def creat_sql_str(cursor=None, table=None, colum=None, value=None):
        if colum is None:
            colum = []
        if value is None:
            value = []
        if table is None:
            table = []
        if cursor is None:
            cursor = []
        str1 = []
        sql_setns = []
        str_temp = []
        filters = ["desc", "time", "login"]
        for idc, _ in enumerate(cursor):
            for idi, i in enumerate(table[idc]):
                for idj, j in enumerate(colum[idc][idi]):
                    for k in value:
                        if i != '' and j != '' and k != '' and j != 'desc' and not any(
                                f in j for f in filters):  # desc is reserve word for sql  2.dont sql time day years
                            if not sql_setns:
                                sql_setns = "select * from " + i + " where " + j + "='" + k + "'"
                            else:
                                sql_setns += " or " + j + "='" + k + "'"
                str_temp.append(idc)
                str_temp.append(idi)
                str_temp.append(sql_setns)
                str1.append(str_temp)
                str_temp = []
                sql_setns = []
        return str1

    def do_sql_from_str(self, str1=None):
        # 0:index cursor =database
        # 1:index table
        # 2:sql str
        if str1 is None:
            str1 = []
        data1 = []
        df_list = []
        for test, i in enumerate(str1):
            if i[2]:
                cursor = self.cursor[i[0]]
                table = self.table[i[0]][i[1]]
                cursor.execute(i[2])
                results_real = cursor.fetchall()
                if results_real != ():  # results_real is tuple so () ,not None or ''
                    data1.append([[self.database[i[0]], table]])
                    col = cursor.description
                    col = self.tuple_2_list(col)
                    df_list.append(pd.DataFrame(list(results_real), columns=col))
        print('done sql in do_sql_from_str')
        with pd.ExcelWriter("result.xlsx") as writer:
            for idd, df in enumerate(df_list):
                df1 = pd.DataFrame(data1[idd], columns=['database', 'table'])
                df2 = df
                out_df = pd.concat([df2, df1]).reset_index(drop=True)
                sheet_index = "sheet" + str(idd)
                out_df.to_excel(writer, sheet_name=sheet_index, index=False)
        self.sheet_union("result.xlsx")
        return df_list

    @staticmethod
    def sheet_union(file_name="result.xlsx"):
        a = pd.read_excel(file_name, sheet_name=None, header=None)
        dfs = []
        for key, value in a.items():
            if key != 'all':
                dfs.append(value)
        df_all = pd.concat(dfs)
        with pd.ExcelWriter(file_name) as writer:
            df_all.to_excel(writer, sheet_name="all", index=False, columns=None)

    @staticmethod
    def tuple_2_list(tuple0=None):
        result = []
        if tuple0 is not None:
            for i in tuple0:
                result.append(i[0])
        return result

    @staticmethod
    def fill_color(file_name="result.xlsx", value1=None):
        if value1 is None:
            value1 = ['40733']
        wb = openpyxl.load_workbook(file_name)
        ws = wb['all']  # Name of the working sheet
        fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')
        fill_cell2 = PatternFill(patternType='solid', fgColor='03FCF4')
        fill_cell3 = PatternFill(patternType='solid', fgColor='35FC03')
        fill_cell4 = PatternFill(patternType='solid', fgColor='FCBA03')

        ws['A1'].fill = fill_cell1
        ws['B1'].fill = fill_cell2
        ws['C1'].fill = fill_cell3
        ws['D1'].fill = fill_cell4

        filter1 = ['database', 'table']
        filter2 = value1

        for i in ws.rows:
            for j in i:
                if any(f in str(j.value) for f in filter1):
                    ws[j.coordinate].fill = fill_cell1
                    for temp in range(j.col_idx):
                        ws.cell(j.row, temp + 1).fill = fill_cell1
                elif any(f in str(j.value) for f in filter2):
                    ws.cell(j.row, j.column).fill = fill_cell2

        wb.save(file_name)
        print('fill color done!')

    def do_quest(self, database_like=None, table_like=None, colum_like=None, value=None):
        if database_like is None:
            database_like = []
        if value is None:
            value = []
        if table_like is None:
            table_like = []
        if colum_like is None:
            colum_like = []
        self.database_like = database_like
        self.table_like = table_like
        self.colum_like = colum_like
        self.value = value
        colum = self.do_colum_like(colum_like)

        self.do_real_worlk(colum, value)
        print('done')
        self.fill_color(file_name="result.xlsx", value1=value)

    def do_colum_like(self, colum_like=None):
        if colum_like is None:
            colum_like = []
        colum = copy.deepcopy(self.colum)
        if colum_like != []:
            for idi, i in enumerate(self.colum):
                for idj, j in enumerate(self.colum[idi]):
                    for idk, k in enumerate(self.colum[idi][idj]):
                        for l in colum_like:
                            if k.find(l) == -1:
                                colum[idi][idj][idk] = ''
                            else:
                                pass
        return colum


def main():
    db = MysqlQuest(host='localhost', user='root', password='root', port=3306)
    db.do_quest(colum_like=[], value=['40733'])


if __name__ == '__main__':
    main()
